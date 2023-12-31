from django.shortcuts import render, redirect, get_object_or_404
from django.http.response import HttpResponse
import io
from PIL import Image
from pdf2image import convert_from_path,convert_from_bytes


# Create your views here.

from django.shortcuts import render
from django.views.generic import FormView
from .forms import *
from django.views.decorators.csrf import csrf_exempt

import pytesseract    # ======= > Add
from PIL import Image
from .forms import UploadFileForm, RotateChoiceForm, OutputChoiceForm

def index(request):
    #return HttpResponse("Hello, world!! at ocr_app's index page.")
    return render(request, 'ocr_app/index.html', {})

def HomeView(request):
    if request.method == 'GET':
        return render(request, 'ocr_app/ocrz_index.html', {
            'form': UploadFileForm(),
            'choiceform': RotateChoiceForm(),
            'outputform': OutputChoiceForm(),
        })
    elif request.method == 'POST':
        form = UploadFileForm(request.POST)
        choiceform = RotateChoiceForm(request.POST)
        outputform = OutputChoiceForm(request.POST)
        if not form.is_valid():
            return render(request, 'ocr_app/ocrz_index.html', {
                'form': form,
                'choiceform': choiceform,
                'outputform': outputform,
            })
        if not choiceform.is_valid():
            return render(request, 'ocr_app/ocrz_index.html', {
                'form': form,
                'choiceform': choiceform,
                'outputform': outputform,
            })
        if not outputform.is_valid():
            return render(request, 'ocr_app/ocrz_index.html', {
                'form': form,
                'choiceform': choiceform,
                'outputform': outputform,
            })
        file = form.cleaned_data['file']
        rotate=choiceform.cleaned_data['angle']
        output=outputform.cleaned_data['output']
        return HttpResponse(rotate,output)   

def PDFZHome(request):
    if request.method == 'GET':
        return render(request, 'ocr_app/pdfz_index.html', {
            'form': UploadFileForm(),
        })
    elif request.method == 'POST':
        form = UploadFileForm(request.POST)
        if not form.is_valid():
            return render(request, 'ocr_app/pdfz_index.html', {
                'form': form,
            })
        file = form.cleaned_data['file']
        return HttpResponse()           

import os
import pyocr
from django.conf import settings

# https://yu-nix.com/archives/django-choicefield/      
from django.shortcuts import render
from django.http import HttpResponse
from .forms import UploadFileForm, RotateChoiceForm, OutputChoiceForm

@csrf_exempt
def process_image(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST)
        choiceform = RotateChoiceForm(request.POST)
        outputform = OutputChoiceForm(request.POST)
        if not choiceform.is_valid():
            return render(request, 'ocr_app/ocrz_index.html', {
                'form': form,
                'choiceform': choiceform,
                'outputform': outputform,
            })
        if not outputform.is_valid():
            return render(request, 'ocr_app/ocrz_index.html', {
                'form': form,
                'choiceform': choiceform,
                'outputform': outputform,
        })
        rotate = choiceform.cleaned_data['angle']
        output=outputform.cleaned_data['output']
        if rotate == 'left':
            angle=90
        elif rotate == 'right':
            angle=-90
        else:
            angle=0
        
        response_data = {}
        upload = request.FILES['file']
        

        # ファイル名の取得
        upload_file_name = upload.name

        # 拡張子を除くファイル名(root)と拡張子(ext)の取得
        root, ext = os.path.splitext(upload_file_name)

        # PDFファイルの場合
        #angle=-90
        #angle=0
        if ext=='.pdf':
            with open(os.path.join(settings.BASE_DIR, 'media/pdf/tmp.pdf'), 'wb+') as f:    #3
                for chunk in upload.chunks():
                    f.write(chunk)
            #pdf_file = open(os.path.join(settings.BASE_DIR, 'media/pdf/tmp.pdf'), 'r')
            org_img=convert_from_path(os.path.join(settings.BASE_DIR, 'media/pdf/tmp.pdf'), fmt='jpg', dpi=200)[0]
            #org_img = convert_from_path('media/pdf/tmp2.pdf', fmt='jpg', dpi=200)[0]
            #org_img2 = convert_from_bytes(open(os.path.join(settings.BASE_DIR, 'media/pdf/tmp.pdf'), 'rb').read())[0]
            #org_img2.save(os.path.join(settings.BASE_DIR, 'media/img/tmp2.jpg'))
            img = org_img.rotate(angle, expand=True)
            img.save(os.path.join(settings.BASE_DIR, 'media/img/tmp.jpg'))
            imgfile = 'media/img/tmp.jpg'

        
        # イメージファイルの場合
        else:
            img=Image.open(upload)
        #'''
        tesseract_path = "/usr/bin"
        if os.name =='nt':
            tesseract_path = r"C:\Program Files\Tesseract-OCR"
        if tesseract_path not in os.environ["PATH"].split(os.pathsep):
            os.environ["PATH"] += os.pathsep+tesseract_path
        pyocr.tesseract.TESSERACT_CMD = '/usr/bin/tesseract'
        if os.name=='nt': 
            pyocr.tesseract.TESSERACT_CMD = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        tools = pyocr.get_available_tools()

        builder = pyocr.builders.LineBoxBuilder(tesseract_layout=6)
        #document = tools[0].image_to_string(img, lang="jpn", builder=builder)
        document = tools[0].image_to_string(img, lang="jpn")
        document=document.replace('、', ',')
        document=document.replace("|", "")
        document=document.replace(", ", ",")
        document=document.replace("],", "1,")
        document=document.replace('],','1,')
        

        import openpyxl as px
        import datetime
        wb = px.Workbook()
        ws = wb.active
        ws['A1'].value = '文字認識結果'
        row = 4
        document = document.split("\n")
        i = 1
        nrow = len(document)
        for text in document:
            text=text.replace('、', ',')
            text=text.replace('l', '')
            text=text.replace(', ', ',')
            text=text.replace('],','1,')
            text=text.replace('|','')
            ws.cell(row=row, column=1).value = row-3
            text =text.split()
            j=2
            for t in text:
                ws.cell(row=row, column = j).value = t
                j=j+1
            row=row+1
        dt =datetime.datetime.now()
        dtime = dt.strftime('%Y%m%d-%H%M%S')
        xlsfile = 'media/excel/ocr-result.xlsx'
        #wb.save(xlsfile)
        
        # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=%s' % 'ocr-result.xlsx'

        # データの書き込みを行なったExcelファイルを保存する
        wb.save(response)
        
        if output=='excel':
            return response

        #return JsonResponse(response_data)
        #return HttpResponse(response_data,'ocrz_list.html')
        #return HttpResponse(document, 'ocrz_list.html')
    
        params={}
        params = {
                'title': '文字認識',
                #'id': upload.id,
                #'setting_form': form,
                #'original_url': upload_image. coverpage.url,
                #'result_url': upload_image.result.url,
                'imgfile': imgfile,            
                'document': document,
                'response': response
        }
        
    return render(request, 'ocr_app/ocrz_ocr.html', params)

#https://awstip.com/django-data-export-generating-csv-and-excel-files-for-user-downloads-a18b14388c9f
import openpyxl
from django.http import HttpResponse

def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

    # Write header row
    header = ['ID', 'Name', 'Email']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = MyModel.objects.all().values_list('id', 'name', 'email')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

from pypdf import PdfReader

def process_pdf(request):

    if request.method == 'POST':
        form = UploadFileForm(request.POST)
    
        upload = request.FILES['file']
        
        # ファイル名の取得
        upload_file_name = upload.name

        # 拡張子を除くファイル名(root)と拡張子(ext)の取得
        root, ext = os.path.splitext(upload_file_name)

        params={}

        # PDFファイルの場合
        if ext=='.pdf':
            with open(os.path.join(settings.BASE_DIR, 'media/pdf/tmp3.pdf'), 'wb+') as f:    #3
                for chunk in upload.chunks():
                    f.write(chunk)
            reader=PdfReader(os.path.join(settings.BASE_DIR, 'media/pdf/tmp3.pdf'))
            page = reader.pages[0]
            document = page.extract_text()

            #params={}
            params = {
                'title': '文字抽出',    
                'document': document,
            }
        return render(request, 'ocr_app/ocrz_pdf.html', params)